"""
Enterprise Report Generator
Creates comprehensive reports in HTML, PDF, and Excel formats
Professional formatting with charts, tables, and executive summaries
"""

from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from pathlib import Path
import json
import base64
from io import BytesIO

# Import advanced HTML generator
try:
    from modules.AdvancedHTMLReportGenerator import AdvancedHTMLReportGenerator
    ADVANCED_HTML_AVAILABLE = True
except ImportError:
    ADVANCED_HTML_AVAILABLE = False

# Report generation libraries
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class EnterpriseReportGenerator:
    """
    Generate professional enterprise reports in multiple formats
    Supports HTML, PDF, and Excel with charts and formatting
    """

    def __init__(self, output_dir: str = "./reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Branding
        self.company_name = "SOC Defense System"
        self.report_title_prefix = "Security Operations Center"

    def generate_html_report(self,
                            report_type: str,
                            data: Dict[str, Any],
                            filename: str = None) -> str:
        """
        Generate professional HTML report

        Args:
            report_type: Type of report
            data: Report data
            filename: Output filename

        Returns:
            Path to generated HTML file
        """
        if not filename:
            filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"

        filepath = self.output_dir / filename

        html_content = self._generate_html_content(report_type, data)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return str(filepath)

    def generate_pdf_report(self,
                           report_type: str,
                           data: Dict[str, Any],
                           filename: str = None) -> str:
        """
        Generate professional PDF report

        Args:
            report_type: Type of report
            data: Report data
            filename: Output filename

        Returns:
            Path to generated PDF file
        """
        if not REPORTLAB_AVAILABLE:
            # Fallback to HTML
            print("ReportLab not available. Generating HTML instead.")
            return self.generate_html_report(report_type, data, filename)

        if not filename:
            filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        filepath = self.output_dir / filename

        doc = SimpleDocTemplate(str(filepath), pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a237e'),
            alignment=TA_CENTER,
            spaceAfter=30
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#283593'),
            spaceAfter=12
        )

        # Title
        title = Paragraph(f"{self.report_title_prefix}<br/>{report_type}", title_style)
        story.append(title)

        # Metadata
        metadata_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Classification: CONFIDENTIAL"
        story.append(Paragraph(metadata_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

        # Executive Summary
        if 'executive_summary' in data:
            story.append(Paragraph("Executive Summary", heading_style))
            for key, value in data['executive_summary'].items():
                story.append(Paragraph(f"<b>{key.replace('_', ' ').title()}:</b> {value}", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))

        # Main content
        for section_name, section_data in data.items():
            if section_name == 'executive_summary':
                continue

            story.append(Paragraph(section_name.replace('_', ' ').title(), heading_style))

            if isinstance(section_data, dict):
                # Create table
                table_data = [[key.replace('_', ' ').title(), str(value)]
                             for key, value in section_data.items()]

                # Skip empty tables
                if not table_data:
                    continue

                table = Table(table_data, colWidths=[3*inch, 3*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey)
                ]))
                story.append(table)

            elif isinstance(section_data, list):
                for item in section_data[:10]:  # Limit to 10 items
                    story.append(Paragraph(f"• {item}", styles['Normal']))

            story.append(Spacer(1, 0.2*inch))

        # Build PDF
        doc.build(story)

        return str(filepath)

    def generate_excel_report(self,
                             report_type: str,
                             data: Dict[str, Any],
                             filename: str = None) -> str:
        """
        Generate professional Excel report with charts

        Args:
            report_type: Type of report
            data: Report data
            filename: Output filename

        Returns:
            Path to generated Excel file
        """
        if not OPENPYXL_AVAILABLE:
            print("openpyxl not available. Generating HTML instead.")
            return self.generate_html_report(report_type, data, filename)

        if not filename:
            filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = self.output_dir / filename

        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Summary sheet
        ws_summary = wb.create_sheet("Executive Summary")
        self._create_summary_sheet(ws_summary, report_type, data)

        # Detail sheets
        for section_name, section_data in data.items():
            if section_name == 'executive_summary':
                continue

            sheet_name = section_name.replace('_', ' ').title()[:31]  # Excel limit
            ws = wb.create_sheet(sheet_name)

            if isinstance(section_data, dict):
                self._create_dict_sheet(ws, section_data)
            elif isinstance(section_data, list):
                self._create_list_sheet(ws, section_data)

        # Save
        wb.save(filepath)

        return str(filepath)

    def generate_compliance_report_bundle(self,
                                         framework: str,
                                         compliance_data: Dict[str, Any]) -> Dict[str, str]:
        """
        Generate complete compliance report bundle in all formats

        Args:
            framework: Compliance framework (SOC2, GDPR, etc.)
            compliance_data: Compliance data

        Returns:
            Dictionary of format -> file path
        """
        report_type = f"{framework}_Compliance_Report"

        return {
            'html': self.generate_html_report(report_type, compliance_data),
            'pdf': self.generate_pdf_report(report_type, compliance_data),
            'excel': self.generate_excel_report(report_type, compliance_data)
        }

    def generate_threat_intelligence_report(self,
                                           threat_data: Dict[str, Any],
                                           format: str = 'html') -> str:
        """
        Generate threat intelligence report

        Args:
            threat_data: Threat intelligence data
            format: Output format (html, pdf, excel)

        Returns:
            Path to generated report
        """
        report_type = "Threat_Intelligence_Report"

        if format == 'html':
            return self.generate_html_report(report_type, threat_data)
        elif format == 'pdf':
            return self.generate_pdf_report(report_type, threat_data)
        elif format == 'excel':
            return self.generate_excel_report(report_type, threat_data)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def generate_incident_report(self,
                                incident_data: Dict[str, Any],
                                format: str = 'pdf') -> str:
        """
        Generate incident investigation report

        Args:
            incident_data: Incident data
            format: Output format

        Returns:
            Path to generated report
        """
        report_type = f"Incident_Report_{incident_data.get('incident_id', 'Unknown')}"

        if format == 'html':
            return self.generate_html_report(report_type, incident_data)
        elif format == 'pdf':
            return self.generate_pdf_report(report_type, incident_data)
        elif format == 'excel':
            return self.generate_excel_report(report_type, incident_data)

    def _generate_html_content(self, report_type: str, data: Dict[str, Any]) -> str:
        """Generate HTML content with professional styling - Uses ADVANCED version with charts and filters"""
        # Try to use advanced HTML generator first
        if ADVANCED_HTML_AVAILABLE:
            try:
                advanced_gen = AdvancedHTMLReportGenerator(self.company_name)
                return advanced_gen.generate_advanced_html(report_type, data)
            except Exception as e:
                print(f"[WARN] Advanced HTML generation failed: {e}. Falling back to basic HTML.")

        # Fallback to basic HTML
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_type}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333;
            background: #f5f5f5;
        }}

        .container {{
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background: white;
        }}

        header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px 20px;
            text-align: center;
            border-radius: 8px 8px 0 0;
        }}

        h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
        }}

        .metadata {{
            background: #f8f9fa;
            padding: 15px;
            border-left: 4px solid #667eea;
            margin: 20px 0;
        }}

        .section {{
            margin: 30px 0;
            padding: 20px;
            border: 1px solid #e0e0e0;
            border-radius: 8px;
        }}

        .section-title {{
            font-size: 1.8em;
            color: #667eea;
            margin-bottom: 15px;
            padding-bottom: 10px;
            border-bottom: 2px solid #667eea;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
        }}

        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}

        th {{
            background: #667eea;
            color: white;
            font-weight: 600;
        }}

        tr:hover {{
            background: #f5f5f5;
        }}

        .kpi-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}

        .kpi-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }}

        .kpi-value {{
            font-size: 2em;
            font-weight: bold;
        }}

        .kpi-label {{
            font-size: 0.9em;
            opacity: 0.9;
            margin-top: 5px;
        }}

        .footer {{
            margin-top: 40px;
            padding: 20px;
            background: #f8f9fa;
            text-align: center;
            color: #666;
            border-radius: 0 0 8px 8px;
        }}

        .watermark {{
            position: fixed;
            bottom: 20px;
            right: 20px;
            opacity: 0.1;
            font-size: 4em;
            color: #667eea;
            transform: rotate(-45deg);
            pointer-events: none;
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>{self.report_title_prefix}</h1>
            <h2>{report_type.replace('_', ' ')}</h2>
        </header>

        <div class="metadata">
            <strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br>
            <strong>Classification:</strong> CONFIDENTIAL<br>
            <strong>Distribution:</strong> Internal Use Only
        </div>
"""

        # Executive Summary with KPIs
        if 'executive_summary' in data:
            html += """
        <div class="section">
            <h2 class="section-title">Executive Summary</h2>
            <div class="kpi-grid">
"""
            for key, value in data['executive_summary'].items():
                html += f"""
                <div class="kpi-card">
                    <div class="kpi-value">{value}</div>
                    <div class="kpi-label">{key.replace('_', ' ').title()}</div>
                </div>
"""
            html += """
            </div>
        </div>
"""

        # Sections
        for section_name, section_data in data.items():
            if section_name == 'executive_summary':
                continue

            html += f"""
        <div class="section">
            <h2 class="section-title">{section_name.replace('_', ' ').title()}</h2>
"""

            if isinstance(section_data, dict):
                html += "<table><thead><tr><th>Metric</th><th>Value</th></tr></thead><tbody>"
                for key, value in section_data.items():
                    html += f"<tr><td>{key.replace('_', ' ').title()}</td><td>{value}</td></tr>"
                html += "</tbody></table>"

            elif isinstance(section_data, list):
                html += "<ul>"
                for item in section_data:
                    html += f"<li>{item}</li>"
                html += "</ul>"

            html += """
        </div>
"""

        html += f"""
        <div class="footer">
            <p>&copy; {datetime.now().year} {self.company_name}. All rights reserved.</p>
            <p>This report contains confidential information. Unauthorized distribution is prohibited.</p>
        </div>
    </div>

    <div class="watermark">CONFIDENTIAL</div>
</body>
</html>
"""

        return html

    def _create_summary_sheet(self, ws, report_type: str, data: Dict[str, Any]):
        """Create Excel summary sheet with formatting"""
        # Title
        ws['A1'] = self.report_title_prefix
        ws['A1'].font = Font(size=20, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 40

        # Report type
        ws['A2'] = report_type.replace('_', ' ')
        ws['A2'].font = Font(size=14, bold=True)
        ws.merge_cells('A2:D2')

        # Metadata
        row = 4
        ws[f'A{row}'] = "Generated:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)

        # Executive summary
        if 'executive_summary' in data:
            row += 2
            ws[f'A{row}'] = "Executive Summary"
            ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            ws.merge_cells(f'A{row}:D{row}')

            row += 1
            for key, value in data['executive_summary'].items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                # Convert complex types to string for Excel
                if isinstance(value, (dict, list)):
                    ws[f'B{row}'] = str(value)
                else:
                    ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _create_dict_sheet(self, ws, data: Dict[str, Any]):
        """Create Excel sheet from dictionary"""

        # Special handling for IoCs with IP intelligence table
        if 'ip_intelligence_table' in data and isinstance(data['ip_intelligence_table'], list) and len(data['ip_intelligence_table']) > 0:
            self._create_ip_intelligence_sheet(ws, data['ip_intelligence_table'])
            return

        # Headers
        ws['A1'] = "Metric"
        ws['B1'] = "Value"

        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")

        # Data
        row = 2
        for key, value in data.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = str(value)
            row += 1

        # Formatting
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

    def _create_ip_intelligence_sheet(self, ws, ip_table: List[Dict[str, Any]]):
        """Create formatted IP intelligence table in Excel"""
        if not ip_table or len(ip_table) == 0:
            return

        # Get column headers from first row
        headers = list(ip_table[0].keys())

        # Set column widths
        column_widths = {
            "IP Address": 18,
            "BAD IP": 12,
            "Risk Score": 12,
            "Risk Level": 12,
            "Attack Count": 14,
            "Country": 15,
            "City": 15,
            "First Seen": 20,
            "Last Seen": 20,
            "Attack Types": 40,
            "Is Whitelisted": 15,
            "Abuse Confidence Score": 20,
            "Total Reports": 15,
            "SANS Count": 12,
            "SANS Attacks": 15
        }

        # Write headers with professional blue color
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")  # Blue 800
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Set column width
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = column_widths.get(header, 15)

        # Write data rows
        for row_idx, ip_data in enumerate(ip_table, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = ip_data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Color code BAD IP column
                if header == 'BAD IP':
                    bad_ip_value = str(value).upper()
                    if bad_ip_value == 'YES':
                        cell.fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")  # Red
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif bad_ip_value == 'NO':
                        cell.fill = PatternFill(start_color="22c55e", end_color="22c55e", fill_type="solid")  # Green
                        cell.font = Font(bold=True, color="000000")
                    elif bad_ip_value == 'N/A':
                        cell.fill = PatternFill(start_color="9ca3af", end_color="9ca3af", fill_type="solid")  # Gray
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif bad_ip_value == 'UNKNOWN':
                        cell.fill = PatternFill(start_color="fbbf24", end_color="fbbf24", fill_type="solid")  # Yellow
                        cell.font = Font(bold=True, color="000000")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Color code by risk level
                risk_level = ip_data.get('Risk Level', '')
                if header == 'Risk Level':
                    if risk_level == 'CRITICAL':
                        cell.fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif risk_level == 'HIGH':
                        cell.fill = PatternFill(start_color="ea580c", end_color="ea580c", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif risk_level == 'MEDIUM':
                        cell.fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
                        cell.font = Font(bold=True, color="000000")
                    elif risk_level == 'LOW':
                        cell.fill = PatternFill(start_color="22c55e", end_color="22c55e", fill_type="solid")
                        cell.font = Font(bold=True, color="000000")

                # Center alignment for certain columns
                if header in ['BAD IP', 'Risk Score', 'Risk Level', 'Attack Count']:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center")

                # Add borders
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                cell.border = thin_border

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

    def _create_list_sheet(self, ws, data: List[Any]):
        """Create Excel sheet from list"""
        # Header
        ws['A1'] = "Item"
        ws['A1'].font = Font(bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")

        # Data
        for idx, item in enumerate(data, start=2):
            ws[f'A{idx}'] = str(item)

        ws.column_dimensions['A'].width = 50
